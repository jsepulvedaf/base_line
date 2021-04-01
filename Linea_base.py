

# -*- coding: utf-8 -*-
#-*- coding: utf-8 -*-
"""
Created on Tue Oct  6 07:07:02 2020

@author: JSEPULVEDA-PC
"""



import streamlit as st
import datetime
from datetime import date
from datetime import datetime
from datetime import timedelta
import pandas as pd
import plotly
import plotly.express as px
from typing import Dict
import openpyxl
from openpyxl import Workbook
from openpyxl.chart import (
    LineChart,
    Reference,
)
from openpyxl.chart.axis import DateAxis
from typing import Dict    
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

import os
from os.path import basename
from os import path
import pandas_profiling as pf
from pandas_profiling import ProfileReport
from streamlit_pandas_profiling import st_profile_report
from openpyxl.writer.excel import ExcelWriter






def main():

    
            
      
        
        
       
                
        ############ingresa el nomre del archivo nombre del archivo ############
        
        st.image('logo consorcio.png', width=85)
        st.title("ESTA APLICACION CALCULA LAS PERDIDAS FISICAS  A 24 HORAS ")
        st.text("by @jsepulvedaf")
        st.subheader("Por  favor subir el archivo teniendo en cuenta que las columnas deben tener el siguiente orden y los encabezados asi:  fecha, Caudal PE PS PC PM")
        
        data_file = st.file_uploader("Suba el archivo",type=['xlsx'])
        datos1= pd.read_excel (data_file)
        datos1['fecha']=pd.to_datetime(datos1['fecha'])
       
        c1, c2,c3 = st.beta_columns((4,4,4))
        with c1:
        
            sec =st.text_input("Entre el Sector")
            sector=sec.upper()     
     
        with c2:
        
          user=st.number_input("entre nuemero de usuarios:")
          km=st.number_input("entre Longitud red km:" )       
          qn=st.number_input("Valor de Consumo nocturno l/s:")
        
          qnls=round(float(qn)*(1/3600)*float(user),2)
        
        with c3:
            grand_c=st.number_input("Consumos nocturno grandes l/s")
        
            n1=st.text_input("entre el N1:")
            n1_favad=float(n1)
        
        nombre_columna = datos1.columns.tolist()    
        seleccion=st.multiselect("seleeciones los de presion a usar", nombre_columna)
        selected = datos1[seleccion]
        selected['P_media']=selected.mean(axis=1)
        Analisis_p = pd.concat([datos1['fecha'], selected], axis=1)
        
        
        
        ########Leer los datos del archivo  _csv#########
        
        
        
        
        
        ########## Calcula la presion Media  #####################
        
           
        
         
        ########une los campos fecha y hora y crea nuevo campo fecha a########
        
        Analisis_p['fecha'] = pd.to_datetime(datos1['fecha'])
       
        datos =pd.concat([datos1['Caudal'],Analisis_p[:]],axis=1, join="inner") 
        # st.dataframe(datos)
        maximo= max(datos['fecha'])
        minimo = min(datos['fecha'])
        pr = ProfileReport(datos, explorative=True)

        st.title("Pandas Profiling in Streamlit")
        st.write(datos)
        st_profile_report(pr)
        
        # ############convierte la fecha en el index#################
        
        datos.set_index('fecha', inplace=True)
        
       
        # ########### Agrega al Dataframe Dia, hora , minutos #########
        datos['Año'] = datos.index.year                                
        datos['Mes'] = datos.index.month                               
        datos['Dia'] = datos.index.day
        datos['hora'] = datos.index.hour                       
        datos['minuto']=datos.index.minute
        #print (datos.head(100))
        # st.dataframe(datos)
        
        hora = str(maximo)
        hora.split()
        hora_str= str(hora[0:10])
        
        # ############  Calcula Qmin promoedio de los 7 dias#############
        P_table=pd.pivot_table(datos,index=['fecha'],values=['Dia','hora','Caudal'])
       
       
        
        # ########### Calcula Q y P medios diarios ##########################
        
        Q_P_med_dia=pd.pivot_table(datos,index=['fecha'],values=['Dia','hora','P_media','Caudal'])
        Q_P_med_dia.rename(columns={'Caudal': 'QMD','P_media':'PMD'}, inplace=True) # renombra cuadal con QMD y pmedia con PMD
        # Q_P_promedio_7= Q_P_med_dia. groupby(['hora'])[['QMD']].mean()
        filtro_hr=  ( Q_P_med_dia['hora'] >=0) & (  Q_P_med_dia['hora']<=5) &( Q_P_med_dia['hora'] >=20) & (  Q_P_med_dia['hora']<=23) 
        filtro_fin=datos.loc[filtro_hr,['Dia','hora','Caudal','P_media']]
        # st.dataframe(filtro_fin)
       
        # ###########  filtra dataframe en las horas de 2 a 3 am #####################
        filter_hr=  (datos['hora'] >= 2) & (datos['hora']<4) 
        
        
        # fiter= (datos['hora']>=2) & (datos['hora']<=3)  #filtro valores de 2 a 3 am
 
 
        
        
        # ############# Crea nuevo dataframe con los datos  filtrados ###############
        filtro=datos.loc[filter_hr,['Dia','hora','P_media','Caudal']]
        
        # filtro['P_prom']=filtro.loc[:,['P_media']].mean(axis=1)
        # fitro['Qmn_prom']=filtro.loc[['Caudal'],:].mean(axis=0)
        # st.dataframe(filtro)
        # ########Calcula el valor de los Q  de las 2 y 3  de minimo consumo ########
        
        Q_table_min=round(pd.pivot_table(filtro,index=['Dia'],values=('Caudal'),aggfunc="min"),2)
        P_table_hmin=pd.pivot_table(filtro,index=['Dia'],values=('P_media'),aggfunc="mean") #  la presion maxima a la hora del mn consumo
        # st.write("Q minimo medio")
        # st.dataframe(Q_table_min)
       
        # ########Calcula el qmin promedio de2 - 4  #########################
        
        Q_min_med= round(Q_table_min.groupby(['Dia'])[['Caudal']].min(),2)
        Q_min_med.rename(columns={'Caudal': 'Qmin_prom'}, inplace=True) #rename 
        Q_min_min = Q_min_med['Qmin_prom'].mean()
        
        # st.write("Q minimo minimo medio")
        # st.write(Q_min_min)
        # ######### escoje el Qmin entre las 2 y las 3 ################3
      
        Q_min=round(Q_table_min.groupby(['Dia'])[['Caudal']].min(),2)
        Q_min.rename(columns={'Caudal': 'Qmin'}, inplace=True)   #rename
        
     
        # ########Calcula el P promedio de2 y 3  #########################
        P_hmin=round(P_table_hmin.groupby(['Dia'])[['P_media']].mean(),2)
        
        
        # #########calcula el Q promedio de los caudales de 2 y 3  en los 7 dias##########
        Q_sector_min_med=round(Q_min_med.groupby('Dia')[['Qmin_prom']].mean(),2)
       
        # ####### Calcula el Qmim promedio de los minimos  en los 7 dias ##############
        
        Q_sector_min=round(float(Q_sector_min_med['Qmin_prom'].mean()),2)
        P_sector_min=round(float(P_hmin['P_media'].mean()),2)
        # P_sector_min.rename(columns={'P_media': 'P_Qmn'}, inplace=True)   #rename
       
        
        st.write("Comportamiento QMN - (lps) / Sector: "+ sector)
        chart = st.line_chart(P_table['Caudal'])
        st.write("ok")
            
        st.header("Resultados")    
        c4, c5,c6,c7 = st.beta_columns((1,1,1,1))
        
        with c4:
            st.write("Fecha de incio")
            st.write(minimo)
            st.write("fecha Final")
            st.write(maximo)
            Num_hab=round(4*user,2)
            st.write("Numero de habitantes")
            st.write(Num_hab)           
           
            Dens_cnx=round( float(user/km),2)
            st.write("DEsnidad de Conexiones")
            st.write(Dens_cnx)
            
        with c5:
           
            p_dia_med=Q_P_med_dia['PMD'].mean()
            st.write("Presion Media dia")
            st.write(round(p_dia_med,2))
                                    
            st.write("Presion minima Noche")
            st.write(round(P_sector_min,2))
            
            hdf=round(pow((p_dia_med/ P_sector_min), n1_favad),2)
            st.write("Factor Hora Dia  ")
            st.write(hdf)
            
        with c6:           
            st.write("Consumo Normal")
            st.write(qnls)
            
            ubl= round(((20*km+1.25*user)*pow((P_sector_min/50),1.5))/3600,2)
            st.write("Ubl")
            st.write(ubl)
            
            UARL=round(((18*km)+(user*(0.8)))*((p_dia_med)/86400),2)
            # UARL24=round(UARL* hdf,2)
            st.write("UARL")
            st.write(UARL)
            
            DNL=Q_sector_min-UARL-qnls- grand_c
            ILI=round(DNL/UARL,2)
            st.write("ILI")
            st.write(ILI)
            
        with c7:      
           
            st.write("Caudal Nocturno Detectable ")
            st.write(DNL)
        
            st.write("MNF ")
            st.write(Q_sector_min)
            
            
            
            
          
            PL24h=round((DNL+UARL)*hdf,2)
            st.write("Perdias fisicas en 24 horas ")
            st.write(PL24h)
           
          
        
            
        # writer = pd.ExcelWriter('archivo_linea base'+ hora_str +'.xlsx')
        # df_final.to_excel(writer, sheet_name="data", index=True)
        # writer.save()
        # writer.close()
        archivo='archivo_plantilla resultados.xlsx'
        wb =load_workbook(filename =archivo )   
        ws = wb.active
       
        
             
  

        ws['E5']= sector
        ws['N4']= minimo
        ws['Q4']= maximo        
        ws['E9'] = km
        ws['M9'] = user
        ws['G15'] = p_dia_med
        ws['Q15'] = P_sector_min
        ws['D17'] = n1_favad
        ws['G17'] = Q_sector_min
        ws['L17'] = ILI
        
        
        ws['O17'] = UARL
        ws['G20'] = qnls+grand_c
        ws['G22'] = Q_sector_min-(qnls+grand_c)
        ws['S19'] = grand_c
        ws['S20'] = qnls
        ws['S23'] = ubl
        ws['R27'] = PL24h
        ws['S17'] = hdf
        
       
        ws1= wb.create_sheet('data') 
        ws1= wb['data']
        with ExcelWriter(archivo) as writer:

               datos.to_excel(writer, sheet_name='data')
       
        now = datetime.now()
        timestamp = datetime.timestamp(now)
       
        hora = str(datetime.fromtimestamp(timestamp))
        hora.split()
        hora_str= str(hora[11:13])+"-"+str(hora[14:16])+"-"+str(hora[17:19])
        file= "Evaluacion_QMN_"+sector+"_"+"_"+hora_str+".xlsx"                          
        
               
        
        wb.save(file)
       
       
                
        
        
if __name__ == '__main__':
    	main()    