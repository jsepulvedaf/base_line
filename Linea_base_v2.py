

# -*- coding: utf-8 -*-
#-*- coding: utf-8 -*-
"""
Created on Tue Oct  6 07:07:02 2020

@author: JSEPULVEDA-PC
"""



import streamlit as st
import datetime
from datetime import date
from datetime import datetime
from datetime import timedelta
import pandas as pd

from typing import Dict

from openpyxl import Workbook
from openpyxl.chart import (
    LineChart,
    Reference,
)
from openpyxl.chart.axis import DateAxis
from typing import Dict    
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import base64
import os
from os.path import basename
from os import path
import io
from plotly.graph_objs import *
import plotly.express as px

# import pandas_profiling as pf

from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.writer.excel import ExcelWriter



rad=st.sidebar.radio('Tipo Linea Base', ['Reparacion Fugas', 'Gestion presion'])


def main():

    
        def to_excel(file_in):
            file_out=BytesIO(file_in)
            procesed=file_out.getvalue()
            return procesed
      
        def tabla_bajar(file_in):
            val = to_excel(file_in)
            b64= base64.b64encode(val)
            return f'<a href= "data:application/octet-stream;base64,{b64.decode()}" download="extract.xlsx">Download csv file</a>'
        
        
               
           
        ############ingresa el nomre del archivo nombre del archivo ############
        
        # st.image('logo consorcio.png', width=85)
        html_temp0=""" <div style= background-color:;padding: 10 px><h2> ESTA APLICACION CALCULA LAS PERDIDAS FISICAS  A 24 HORAS </h2> </div>"""
        st.markdown(html_temp0,unsafe_allow_html=True)  
        #st.warning("ESTA APLICACION CALCULA LAS PERDIDAS FISICAS  A 24 HORAS ")
        #st.text("")
        st.text("by @jsepulvedaf")
        st.info("Por  favor subir el archivo teniendo en cuenta que las columnas deben tener el siguiente orden y los encabezados asi:  fecha, Caudal PE PS PC PM")
        
        data_file = st.file_uploader("Suba el archivo",type=['xlsx'])
        datos= pd.read_excel (data_file)
        datos['fecha']=pd.to_datetime(datos['fecha'])
        datos['new_date'] = [d.date() for d in datos['fecha']]
        html_temp1=""" <div style= background-color:#edb06b;padding: 10 px><h4> DATOS INICIALES </h4> </div>"""
        st.markdown(html_temp1,unsafe_allow_html=True)    
        my_expander = st.expander("")
        with my_expander:
        
        
           
            
            
            
                     
            
            c1, c2,c3 = st.columns((4,4,4))
                 
            
           
            with c1:
            
                sec =st.text_input("Entre el Sector")
                sector=sec.upper()     
         
            with c2:
            
              usuarios_n=st.text_input("entre nuemero de usuarios:")
              user=float(usuarios_n)
              km_t=st.text_input("entre Longitud red km:" )
              km= float(km_t)
              qn_t=st.text_input("Valor de Consumo nocturno l/s:")
              qn=float(qn_t)
            
              qnls=round(qn*(1/3600)*user,2)
            
            with c3:
                
                grande_c=st.text_input("Consumos nocturno grandes l/s")
                grand_c=float(grande_c)
                n1=st.text_input("entre el N1:")
                n1_favad=float(n1)
        
            nombre_columna = datos.columns.tolist()    
            seleccion=st.multiselect("Seleeciones las presiones ", nombre_columna)
            selected = datos[seleccion]
            # st.dataframe(selected)
            selected['P_media']=selected.mean(axis=1)
            Analisis_p = pd.concat([datos['fecha'], selected], axis=1)
            datos['P_media']=Analisis_p['P_media']
        
        ########Leer los datos del archivo  _csv#########
        
        
        
        
        
        ########## Calcula la presion Media  #####################
        
           
        
         
        ########une los campos fecha y hora y crea nuevo campo fecha a########
        
        Analisis_p['fecha'] = pd.to_datetime(datos['fecha'])
        Analisis_p['new_date'] = pd.to_datetime(datos['new_date'])
        # datos =pd.concat([datos,Analisis_p[:]],axis=1, join="inner") 
        # st.dataframe(datos)
        maximo= max(datos['new_date'])
        minimo = min(datos['new_date'])
        # pr = ProfileReport(datos, explorative=True)

        # st.title("Pandas Profiling in Streamlit")
        # st.write(datos)
        # st_profile_report(pr)
        
        # ############convierte la fecha en el index#################
        
        datos.set_index('fecha', inplace=True)
        
       
        # ########### Agrega al Dataframe Dia, hora , minutos #########
        datos['Año'] = datos.index.year                                
        datos['Mes'] = datos.index.month                               
        datos['Dia'] = datos.index.day
        datos['hora'] = datos.index.hour                       
        datos['minuto']=datos.index.minute
        #print (datos.head(100))
        st.dataframe(datos)
        
        hora = str(maximo)
        hora.split()
        hora_str= str(hora[0:10])
        
        # ############  Calcula Qmin promoedio de los 7 dias#############
        P_table=pd.pivot_table(datos,index=['fecha'],values=['Dia','hora','Caudal'])
       
        nom_col = datos.columns.tolist()    
        selec_1=st.multiselect("Seleeciones las presiones ", nom_col)
        selected1 = datos[selec_1]            
        
        
        # ########### Calcula Q y P medios diarios ##########################
        
        Q_P_med_dia=pd.pivot_table(datos,index=['fecha'],values=['Dia','hora','P_media','Caudal','PS','PC'])
        Q_P_med_dia.rename(columns={'Caudal': 'QMD','P_media':'PMD'}, inplace=True) # renombra cuadal con QMD y pmedia con PMD
        # Q_P_promedio_7= Q_P_med_dia. groupby(['hora'])[['QMD']].mean()
        filtro_hr=  ( Q_P_med_dia['hora'] >=0) & ( Q_P_med_dia['hora']<=5) &(Q_P_med_dia['hora'] >=20) & (Q_P_med_dia['hora']<=23) 
        filtro_fin=datos.loc[:,['Dia','hora','Caudal','P_media','PS','PC']]
        Q_table_mean=round(pd.pivot_table(filtro_fin,index=['Dia'],values=('Caudal'),aggfunc="mean"),2)
        P_table_hmean=pd.pivot_table(filtro_fin,index=['Dia'],values=['P_media','PS','PC'],aggfunc="mean") 
        P_table_hmean.rename(columns={'PC': 'PC_D','PS':'PS_D','P_media':'PMD'}, inplace=True) 
        Q_table_mean['Volumen']=round((Q_table_mean['Caudal']/1000)*86400,0)
        
        # resumenPyQ= pd.merge(Q_table_mean,P_table_hmean,on='Dia',how='inner')
        
        fecha_7=datos.groupby(['new_date'])[['Caudal','Dia']].mean()
        
        fecha_7['new_date']= fecha_7.index
        fecha_7.set_index('Dia',inplace=True)
        # st.dataframe(filtro_fin)

# ###########  filtra dataframe en las horas de 2 a 4 am #####################
        filter_hr=  (datos['hora'] >= 2) & (datos['hora']<4) 


       
        
        # ############# Crea nuevo dataframe con los datos  filtrados ###############
        filtro=datos.loc[filter_hr,['Dia','hora','P_media','Caudal','PS','PC']]

# filtro['P_prom']=filtro.loc[:,['P_media']].mean(axis=1)
# fitro['Qmn_prom']=filtro.loc[['Caudal'],:].mean(axis=0)
# st.dataframe(filtro)
# ########Calcula el valor de los Q  de las 2 y 3  de minimo consumo ########

        Q_table_min=round(pd.pivot_table(filtro,index=['Dia'],values=('Caudal'),aggfunc="min"),2)
        Q_table_min.rename(columns={'Çaudal':'Qmin'}, inplace=True)
        P_table_hmin=pd.pivot_table(filtro,index=['Dia'],values=['P_media','PS','PC'],aggfunc="mean") #  la presion maxima a la hora del mn consumo
        
        P_table_hmin.rename(columns={'PC': 'PC_N','PS':'PS_N'}, inplace=True) 
        
        resumenPyQ= pd.merge(Q_table_mean,P_table_hmean,on='Dia',how='inner')
        
        Resumen_tot=pd.merge(Q_table_min,resumenPyQ,on='Dia',how='inner')
        
        
        
        Presiones_DN=pd.concat([fecha_7['new_date'],P_table_hmean,P_table_hmin],axis=1)
        
        # print("Q minimo medio")
        #st.dataframe(Q_table_min)
        
        # ########Calcula el qmin promedio de2 - 4  #########################
        
        Q_min_med= round(Q_table_min.groupby(['Dia'])[['Caudal']].min(),2)
        Q_min_med.rename(columns={'Caudal': 'Qmin_prom'}, inplace=True) #rename 
        Q_min_min = Q_min_med['Qmin_prom'].mean()
        
        # print("Q minimo minimo medio")
         #print(Q_min_min)
         # ######### escoje el Qmin entre las 2 y las 3 ################3
          
        Q_min=round(Q_table_min.groupby(['Dia'])[['Caudal']].min(),2)
        Q_min.rename(columns={'Caudal': 'Qmin'}, inplace=True)   #rename
        
         
        # ########Calcula el P promedio de2 y 3  #########################
        P_hmin=round(P_table_hmin.groupby(['Dia'])[['P_media']].mean(),2)
        
        
        # #########calcula el Q promedio de los caudales de 2 y 3  en los 7 dias##########
        Q_sector_min_med=round(Q_min_med.groupby('Dia')[['Qmin_prom']].mean(),2)
        
        # ####### Calcula el Qmim promedio de los minimos  en los 7 dias ##############
        
        Q_sector_min=round(float(Q_sector_min_med['Qmin_prom'].mean()),2)
        P_sector_min=round(float(P_hmin['P_media'].mean()),2)
        # P_sector_min.rename(columns={'P_media': 'P_Qmn'}, inplace=True)   #rename

  

       
        html_temp3=""" <div style= background-color:#abdbe3;padding: 10px><h4> GRAFICOS </h4> </div>"""
        st.markdown(html_temp3,unsafe_allow_html=True)    
        my_expander1 = st.expander("")
        with my_expander1:
            
           
        
            H1=("Comportamiento QMN - (lps) / Sector: "+ sector)
            fig = px.line(datos['Caudal'], title = H1 )  
            #fig.show()
            # fig.write_html("presiones cali.html")
        
         
            
            
            st.plotly_chart(fig, use_container_width=True)
    
            H2=("Comportamiento Presiones - (m.c.a) / Sector: "+ sector)
            fig2 = px.line(selected1, title = H2 )  
            #fig.show()
            st.plotly_chart(fig2, use_container_width=True)
            # fig.write_html("presiones cali.html")
        
         
            
            
            #st.plotly_chart(fig, use_container_width=True)
    
        
       
        
       # chart = st.line_chart(P_table['Caudal'])
       #  st.write("ok")
        #st.header("Resultados")    
        html_temp=""" <div style= background-color:#c9ffcb;padding: 10px;  P {color:WHITE;}><h4> RESULTADOS </h4> </div>"""
        st.markdown(html_temp,unsafe_allow_html=True)             
        my_expander2 = st.expander("")
        with my_expander2:
            c4, c5,c6 = st.columns((1,1,1))
        
            with c4:
                #st.markdown("""<style>.big-font {font-size:18px !important;}</style>""", unsafe_allow_html=True)
                #st.markdown('<p class="big-font">Fecha de inicio</p>', unsafe_allow_html=True)
                st.write("Fecha de incio")
                st.success(minimo)
                st.write("fecha Final")
                st.success(maximo)
                Num_hab=round(4*user,2)
                st.write("Num. de habitantes")
                st.success(Num_hab)           
               
                Dens_cnx=round( float(user/km),2)
                st.write("Densidad  Conex.")
                st.success(Dens_cnx)
                
            with c5:
               
                p_dia_med=Q_P_med_dia['PMD'].mean()
                st.write("P. Media dia-m.c.a")
                st.warning(round(p_dia_med,2))
                                        
                st.write("P min. Noche- m.c.a")
                st.warning(round(P_sector_min,2))
                
                hdf=round(pow((p_dia_med/ P_sector_min), n1_favad),2)
                st.write("Factor Hora Dia  ")
                st.warning(hdf)
                
            with c6:           
                st.write("Q Noct. Normal-lps")
                st.info(qnls)
                
                ubl= round(((20*km+1.25*user)*pow((P_sector_min/50),1.5))/3600,2)
                st.write("UBL")
                st.info(ubl)
                
                UARL=round(((18*km)+(user*(0.8)))*((p_dia_med)/86400),2)
                # UARL24=round(UARL* hdf,2)
                st.write("UARL")
                st.info(UARL)
                
                # Q_sector_min-qnls-grand_c-ubl*6
                           
                DNL=Q_sector_min-qnls-grand_c-ubl*6  #q FUGAS DECTECTABLES RECUPERABLES POR fUGAS 
                CARL=Q_sector_min-qnls-grand_c
                ILI=round((CARL)/UARL,2)
                st.write("ILI")
                st.info(ILI)
                
        
                
                PERD_FIS=Q_sector_min-qnls-grand_c
                
                               
                DNLP=Q_sector_min-qnls-grand_c  #Q GUFAS RECUPERABLES POR PRESION
                
        html_temp=""" <div style= background-color:#c9ffcb;padding: 10px;  P {color:WHITE;}><h4> CAUDALES </h4> </div>"""
        st.markdown(html_temp,unsafe_allow_html=True)             
        my_expander3 = st.expander("")
        with my_expander3: 
            
            c7, c8 = st.columns((1,1))
            
            with c7:
                st.write("Q MNF-lps")
                st.success(Q_sector_min)
                
                st.write("Q Noct. Detectable-lps")
                st.success(round(DNL,2))
            
                st.write("QF Rec Gestion P")
                st.success(round(DNLP,2))
                        
            with c8:      
                PL24hf=round((DNL)*hdf,2)
                st.write("Q fisicas 24 hr-lps_fugas")
                st.success(PL24hf)
                
              
                PL24hp=round((DNLP)*hdf,2)
                st.write("Q fisicas 24 hr-lps_presion")
                st.success(PL24hp)
               
              
        Dir_res= {'Qmn':Q_sector_min,'Consumo noct. normal':qnls,'Gr Consumos Noc': grand_c, 'ELL':ubl*6,
          'fugas detectables_fugas': DNL, 'Qfugas recuperable_presion': DNLP,'perdida fisica 24h_f': PL24hf,'perdida fisica 24h_p': PL24hp}

        Resumen = pd.DataFrame(data=Dir_res,index=(0,1))
        
            
        # writer = pd.ExcelWriter('archivo_linea base'+ hora_str +'.xlsx')
        # df_final.to_excel(writer, sheet_name="data", index=True)
        # writer.save()
        # writer.close()
        if rad=='Reparacion Fugas':
            #desktop = os.path.normpath(os.path.expanduser("~/Desktop"))     
            archivo='archivo_plantilla resultados_f.xlsx'
            wb =load_workbook(filename =archivo )   
            ws = wb['Calculo QMN']
           
              
      
    
            ws['E5']= sector
            ws['N4']= minimo
            ws['Q4']= maximo        
            ws['E9'] = km
            ws['M9'] = user
            ws['G15'] = p_dia_med
            ws['Q15'] = P_sector_min
            ws['D17'] = n1_favad
            ws['G17'] = Q_sector_min
            ws['L17'] = ILI
            
            
            ws['O17'] = UARL
            ws['G20'] = qnls+grand_c
            ws['G22'] = PERD_FIS
            ws['S19'] = grand_c
            ws['S20'] = qnls
            ws['S23'] = ubl
            ws['R27'] = PL24hf
            ws['S17'] = hdf
            
                    # ruta=os.path.expanduser("~")       
            now = datetime.now()
            timestamp = datetime.timestamp(now)
           
            hora = str(datetime.fromtimestamp(timestamp))
            hora.split()
            hora_str= str(hora[11:13])+"-"+str(hora[14:16])+"-"+str(hora[17:19])
           
    
            ws4 = wb["Presiones"]
            for r in dataframe_to_rows(Presiones_DN, index=True,header=True):
                 ws4.append(r)
            #wb.create_sheet('Data6')
            ws1 = wb["Data6"]
            for r in dataframe_to_rows(Resumen_tot, index=True,header=True):
                 ws1.append(r)
            
            # wb.create_sheet('resumen')
            ws2 = wb["resumen"]
            for r in dataframe_to_rows(Resumen, index=True,header=True):
                 ws2.append(r)
     
        
            wb.create_sheet("graph_p_q")
            ws5 = wb["graph_p_q"]
            for r in dataframe_to_rows(datos, index=True,header=True):
                 ws5.append(r)
        
        
       
            boton=st.button("exportar XLSX")        
                   
           
           
    
            desktop = os.path.normpath(os.path.expanduser("~/Desktop/Resultado"))           
            file= desktop+"\\"+"Evaluacion_QMN_GF_"+sector+"_"+"_"+hora_str+".xlsx" 
            if boton :
                wb.save(file)
                st.write("archivo guardado")
            
            
           
           
            
        if rad=='Gestion presion':
                    #desktop = os.path.normpath(os.path.expanduser("~/Desktop"))     
                    archivo='archivo_plantilla resultados_p.xlsx'
                    wb =load_workbook(filename =archivo )   
                    ws = wb['Calculo QMN']
                   
                      
              
            
                    ws['E5']= sector
                    ws['N4']= minimo
                    ws['Q4']= maximo        
                    ws['E9'] = km
                    ws['M9'] = user
                    ws['G15'] = p_dia_med
                    ws['Q15'] = P_sector_min
                    ws['D17'] = n1_favad
                    ws['G17'] = Q_sector_min
                    ws['L17'] = ILI
                    
                    
                    ws['O17'] = UARL
                    ws['G20'] = qnls+grand_c
                    ws['G22'] = PERD_FIS
                    ws['S19'] = grand_c
                    ws['S20'] = qnls
                    ws['S23'] = ubl
                    ws['R27'] = PL24hp
                    ws['S17'] = hdf
                    
                            # ruta=os.path.expanduser("~")       
                    now = datetime.now()
                    timestamp = datetime.timestamp(now)
                   
                    hora = str(datetime.fromtimestamp(timestamp))
                    hora.split()
                    hora_str= str(hora[11:13])+"-"+str(hora[14:16])+"-"+str(hora[17:19])
                   
            
                    ws4 = wb["Presiones"]
                    for r in dataframe_to_rows(Presiones_DN, index=True,header=True):
                         ws4.append(r)
                    #wb.create_sheet('Data6')
                    ws1 = wb["Data6"]
                    for r in dataframe_to_rows(Resumen_tot, index=True,header=True):
                         ws1.append(r)
                    
                    # wb.create_sheet('resumen')
                    ws2 = wb["resumen"]
                    for r in dataframe_to_rows(Resumen, index=True,header=True):
                         ws2.append(r)
                         
                         
                    wb.create_sheet("graph_p_q")
                    ws5 = wb["graph_p_q"]
                    for r in dataframe_to_rows(datos, index=True,header=True):
                           ws5.append(r)
            
                    boton=st.button("exportar XLSX")   
            
                    desktop = os.path.normpath(os.path.expanduser("~/Desktop/Resultado"))           
                    file= desktop+"\\"+"Evaluacion_QMN_GP_"+sector+"_"+"_"+hora_str+".xlsx" 
                    if boton :
                       wb.save(file)
                       st.write("archivo guardado")
            
                    
                    
                   
                    
                    
                    
                    
                    
                    
        
        
        
        # workbook = openpyxl.load_workbook(file)
        
        # output = io.BytesIO()
        # output.seek(0)
        # libri = output.read()

        # bs64 = base64.b64encode(libri).decode("UTF-8")
                      
          # some strings <-> bytes conversions necessary here
        # linko= f'<a href="data:application/octet-stream;base64,{bs64}" download="file.xlsx">Bajar Archivo</a>'
        # st.markdown(linko, unsafe_allow_html=True)
        
if __name__ == '__main__':
    	main()    